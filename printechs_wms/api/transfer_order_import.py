# -*- coding: utf-8 -*-
"""
Import WMS Transfer Order from Excel (Header + Detail like ASN),
BUT Item sheet uses STORE allocation columns (from column D until first blank header).

Sheets:
  1) "WMS Transfer Order"        -> header rows
  2) "WMS Transfer Order Item"   -> item rows:
        A=parent, B=item_code, C=remarks, D..=STORE columns (header = store code/name)
        Read store columns starting from D until first blank header.
        For each row -> creates one item line per store column where qty > 0.

Parent linking:
  - item.parent can match header.name OR header.to_title
  - recommended to use header.to_title in Excel (human friendly)

Warehouse resolving:
  - Your Warehouse has custom field 'code' (not warehouse_code)
  - Resolver checks: exact name -> code -> warehouse_name -> fallback any non-group

Args:
  file_url: /files/xxx.xlsx
  sheet_name: optional header sheet name (default "WMS Transfer Order")
  create_new: 1 (default) create new (or update safely if existing), 0 update existing by name/to_title
  submit: 0/1

Returns:
  {
    ok, total_items, created, updated, errors, first_to, per_to, sheets_used
  }
"""

from __future__ import annotations

import io
import frappe
from frappe.utils import cstr, flt, getdate
from openpyxl import load_workbook


# ----------------------------
# Sheet names
# ----------------------------
SHEET_HEADER_DEFAULT = "WMS Transfer Order"
SHEET_ITEM_DEFAULT = "WMS Transfer Order Item"

# Item sheet: store columns start at Column D (0-based index 3)
ITEM_STORE_START_COL = 3


# ----------------------------
# Column maps (scrubbed)
# ----------------------------
TO_HEADER_MAP = {
    "name": "name",
    "to_title": "to_title",
    "to title": "to_title",

    "asn": "asn",
    "asn_no": "asn",
    "asn no": "asn",

    "from_warehouse_code": "from_warehouse_code",
    "from warehouse code": "from_warehouse_code",

    "from_warehouse": "from_warehouse",
    "from warehouse": "from_warehouse",

    "prepared_by": "prepared_by",
    "prepared by": "prepared_by",

    "required_date": "required_date",
    "required date": "required_date",
    "required_date_(yyyy_mm_dd)": "required_date",

    "status": "status",
    "remarks": "remarks",
}


def _scrub(s):
    if s is None:
        return ""
    return cstr(s).strip().lower().replace(" ", "_").replace("-", "_")


def _cell_value(cell):
    v = cell.value
    if v is None:
        return None
    # If datetime/date, normalize
    try:
        if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
            return getdate(v)
    except Exception:
        pass
    return v


def _excel_to_rows(file_url=None, fcontent=None, filepath=None, sheet_name=None):
    """Load workbook and return list-of-rows for given sheet."""
    if file_url:
        f = frappe.get_doc("File", {"file_url": file_url})
        path_or_bytes = f.get_full_path()
    elif fcontent:
        path_or_bytes = io.BytesIO(fcontent)
    elif filepath:
        path_or_bytes = filepath
    else:
        return []

    wb = load_workbook(filename=path_or_bytes, data_only=True, read_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        rows = [[_cell_value(c) for c in row] for row in ws.iter_rows()]
        return rows
    finally:
        wb.close()


def _map_row(row, headers, field_map):
    out = {}
    for idx, col in enumerate(headers):
        if idx >= len(row):
            break
        key = _scrub(col)
        if not key:
            continue
        field = field_map.get(key) or key
        val = row[idx]
        if val is not None and (val != "" or isinstance(val, (int, float))):
            out[field] = val
    return out


def _resolve_from_warehouse(from_warehouse_code: str, from_warehouse: str) -> tuple[str, str]:
    """
    Resolve Warehouse for your environment.
    Warehouse fields: code, warehouse_name, name (standard)

    Priority:
      1) from_warehouse is exact Warehouse.name
      2) from_warehouse_code matches Warehouse.code
      3) from_warehouse matches Warehouse.warehouse_name
      4) from_warehouse_code matches Warehouse.name
      5) fallback any non-group warehouse

    Returns: (resolved_code, resolved_warehouse_name)
    """
    code = cstr(from_warehouse_code or "").strip()
    name = cstr(from_warehouse or "").strip()

    # 1) exact Warehouse.name
    if name and frappe.db.exists("Warehouse", name):
        wh_code = frappe.db.get_value("Warehouse", name, "code") or code or name
        return wh_code, name

    # 2) by custom code
    if code:
        wh = frappe.db.get_value("Warehouse", {"code": code}, "name")
        if wh:
            return code, wh

    # 3) by warehouse_name
    if name:
        wh = frappe.db.get_value("Warehouse", {"warehouse_name": name}, "name")
        if wh:
            wh_code = frappe.db.get_value("Warehouse", wh, "code") or code or wh
            return wh_code, wh

    # 4) treat code as name
    if code and frappe.db.exists("Warehouse", code):
        wh_code = frappe.db.get_value("Warehouse", code, "code") or code
        return wh_code, code

    # 5) fallback any non-group
    any_wh = (
        frappe.db.get_value("Warehouse", {"is_group": 0}, "name")
        or frappe.db.get_value("Warehouse", {}, "name")
    )
    if any_wh:
        any_code = frappe.db.get_value("Warehouse", any_wh, "code") or code or any_wh
        return any_code, any_wh

    frappe.throw("No Warehouse found in system.")


def _find_existing_to(header: dict) -> str | None:
    """
    Find existing WMS Transfer Order by:
      1) header.name
      2) header.to_title
    """
    name = cstr(header.get("name") or "").strip()
    to_title = cstr(header.get("to_title") or "").strip()

    if name and frappe.db.exists("WMS Transfer Order", name):
        return name

    if to_title:
        found = frappe.db.get_value("WMS Transfer Order", {"to_title": to_title}, "name")
        if found:
            return found

    return None


def _get_store_columns_from_item_sheet(headers):
    """
    Item sheet format:
      A=parent, B=item_code, C=remarks, D.. = store columns until first blank header
    Returns: list of (col_index, store_name)
    """
    cols = []
    for idx in range(ITEM_STORE_START_COL, len(headers)):
        h = headers[idx]
        store = cstr(h).strip() if h is not None else ""
        if not store:
            break
        cols.append((idx, store))
    return cols


def _build_items_from_store_columns(i_headers, i_data):
    """
    Creates normalized item rows from store allocation columns.
    Output rows format:
      {parent, item_code, store, allocated_qty, remarks, sorted_qty, packed_qty}
    """
    store_cols = _get_store_columns_from_item_sheet(i_headers)
    out = []

    if not store_cols:
        return out

    for row in i_data:
        parent = cstr(row[0] if len(row) > 0 else "").strip()
        item_code = cstr(row[1] if len(row) > 1 else "").strip()
        remarks = cstr(row[2] if len(row) > 2 else "").strip() or None

        if not parent or not item_code:
            continue

        for col_idx, store_name in store_cols:
            if col_idx >= len(row):
                continue
            qty = flt(row[col_idx], 2)
            if qty and qty > 0:
                out.append({
                    "parent": parent,
                    "item_code": item_code,
                    "store": store_name,
                    "allocated_qty": qty,
                    "remarks": remarks,
                    "sorted_qty": 0,
                    "packed_qty": 0,
                })

    return out


@frappe.whitelist()
def import_transfer_order_from_excel(file_url=None, sheet_name=None, create_new=1, submit=0):
    """
    Import Transfer Orders from Excel:
      - header sheet: WMS Transfer Order (or sheet_name)
      - item sheet: WMS Transfer Order Item
    """
    create_new = int(create_new) if create_new is not None else 1
    submit = int(submit) if submit is not None else 0

    errors: list[str] = []
    created: list[str] = []
    updated: list[str] = []
    per_to: list[dict] = []
    first_to = None

    try:
        header_sheet = sheet_name or SHEET_HEADER_DEFAULT

        header_rows = _excel_to_rows(file_url=file_url, sheet_name=header_sheet)
        if not header_rows or len(header_rows) < 2:
            return {
                "ok": False,
                "total_items": 0,
                "created": [],
                "updated": [],
                "errors": [f"Header sheet '{header_sheet}' must have a header row and at least one data row."],
                "first_to": None,
                "per_to": [],
            }

        item_rows = _excel_to_rows(file_url=file_url, sheet_name=SHEET_ITEM_DEFAULT)
        if not item_rows or len(item_rows) < 2:
            return {
                "ok": False,
                "total_items": 0,
                "created": [],
                "updated": [],
                "errors": [f"Item sheet '{SHEET_ITEM_DEFAULT}' must have a header row and at least one data row."],
                "first_to": None,
                "per_to": [],
            }

        h_headers = header_rows[0]
        h_data = header_rows[1:]

        i_headers = item_rows[0]
        i_data = item_rows[1:]

        # ------------------------
        # Parse header rows
        # ------------------------
        header_docs: list[dict] = []
        for row in h_data:
            h = _map_row(row, h_headers, TO_HEADER_MAP)
            if not h:
                continue

            # ignore fully empty rows
            if not any(v is not None and cstr(v).strip() != "" for v in h.values()):
                continue

            rd = h.get("required_date")
            if rd and hasattr(rd, "isoformat"):
                try:
                    h["required_date"] = getdate(rd)
                except Exception:
                    pass

            header_docs.append(h)

        if not header_docs:
            return {
                "ok": False,
                "total_items": 0,
                "created": [],
                "updated": [],
                "errors": ["No header rows found in header sheet."],
                "first_to": None,
                "per_to": [],
            }

        # Index headers by name and to_title
        headers_by_name = {}
        headers_by_title = {}
        for h in header_docs:
            hn = cstr(h.get("name") or "").strip()
            ht = cstr(h.get("to_title") or "").strip()
            if hn:
                headers_by_name[hn] = h
            if ht:
                headers_by_title[ht] = h

        # ------------------------
        # Parse items from store columns
        # ------------------------
        items = _build_items_from_store_columns(i_headers, i_data)

        if not items:
            return {
                "ok": False,
                "total_items": 0,
                "created": [],
                "updated": [],
                "errors": ["No item rows found. Check: parent/item_code and store headers from column D."],
                "first_to": None,
                "per_to": [],
            }

        # Group items by resolved header
        grouped: dict[str, list[dict]] = {}
        for it in items:
            p = cstr(it.get("parent") or "").strip()
            h = headers_by_name.get(p) or headers_by_title.get(p)
            if not h:
                errors.append(f"Item parent '{p}' not found in header sheet (match header.name or header.to_title).")
                continue

            key = cstr(h.get("name") or h.get("to_title") or p).strip()
            grouped.setdefault(key, []).append(it)

        if not grouped:
            return {
                "ok": False,
                "total_items": 0,
                "created": [],
                "updated": [],
                "errors": errors or ["No valid items could be matched to headers."],
                "first_to": None,
                "per_to": [],
            }

        # ------------------------
        # Create/update docs
        # ------------------------
        total_items_imported = 0

        for key, its in grouped.items():
            h = headers_by_name.get(key) or headers_by_title.get(key)
            if not h:
                # fallback scan
                for hh in header_docs:
                    if cstr(hh.get("name") or "").strip() == key or cstr(hh.get("to_title") or "").strip() == key:
                        h = hh
                        break

            if not h:
                errors.append(f"Internal: could not resolve header for key '{key}'.")
                continue

            header_name = cstr(h.get("name") or "").strip()
            to_title = cstr(h.get("to_title") or "").strip()
            asn = cstr(h.get("asn") or "").strip() or None
            status = cstr(h.get("status") or "Draft").strip()
            prepared_by = cstr(h.get("prepared_by") or "").strip() or frappe.session.user
            required_date = h.get("required_date") or None
            remarks = cstr(h.get("remarks") or "").strip() or None

            from_code = cstr(h.get("from_warehouse_code") or "").strip()
            from_wh = cstr(h.get("from_warehouse") or "").strip()
            from_code, from_wh = _resolve_from_warehouse(from_code, from_wh)

            existing = _find_existing_to({"name": header_name, "to_title": to_title})

            do_update = False
            if existing:
                if create_new == 0:
                    do_update = True
                else:
                    # create_new=1 but doc exists -> update to avoid duplicates
                    do_update = True

            if do_update and existing:
                doc = frappe.get_doc("WMS Transfer Order", existing)

                # Replace child table
                doc.items = []

                # Header fields
                doc.from_warehouse_code = from_code
                doc.from_warehouse = from_wh
                if to_title:
                    doc.to_title = to_title
                if asn:
                    doc.asn = asn
                if required_date:
                    doc.required_date = required_date
                if status:
                    doc.status = status
                doc.remarks = remarks
                doc.prepared_by = prepared_by

                for it in its:
                    doc.append("items", {
                        "item_code": it["item_code"],
                        "allocated_qty": flt(it.get("allocated_qty"), 2),
                        "store": (it.get("store") or "").strip(),
                        "sorted_qty": flt(it.get("sorted_qty"), 2) or 0,
                        "packed_qty": flt(it.get("packed_qty"), 2) or 0,
                        "remarks": it.get("remarks"),
                    })

                doc.total_allocated_qty = sum(flt(r.allocated_qty, 2) for r in doc.items)
                doc.items_count = len(doc.items)
                doc.flags.ignore_validate_update_after_submit = True
                doc.save(ignore_permissions=True)

                if submit and int(doc.docstatus) == 0:
                    try:
                        doc.submit()
                    except Exception:
                        errors.append(f"Failed to submit {doc.name}: {cstr(frappe.get_traceback())[:250]}")

                updated.append(doc.name)
                first_to = first_to or doc.name
                per_to.append({"to": doc.name, "items": len(doc.items), "mode": "updated"})
                total_items_imported += len(doc.items)

            else:
                doc = frappe.new_doc("WMS Transfer Order")
                doc.naming_series = getattr(doc, "naming_series", None) or "WMS-TO-.#####"

                doc.to_title = to_title or f"TO-{frappe.utils.now_datetime().strftime('%Y%m%d-%H%M%S')}"
                doc.from_warehouse_code = from_code
                doc.from_warehouse = from_wh
                doc.asn = asn
                doc.required_date = required_date
                doc.status = status
                doc.remarks = remarks
                doc.prepared_by = prepared_by

                for it in its:
                    doc.append("items", {
                        "item_code": it["item_code"],
                        "allocated_qty": flt(it.get("allocated_qty"), 2),
                        "store": (it.get("store") or "").strip(),
                        "sorted_qty": flt(it.get("sorted_qty"), 2) or 0,
                        "packed_qty": flt(it.get("packed_qty"), 2) or 0,
                        "remarks": it.get("remarks"),
                    })

                doc.total_allocated_qty = sum(flt(r.allocated_qty, 2) for r in doc.items)
                doc.items_count = len(doc.items)

                doc.insert(ignore_permissions=True)

                if submit:
                    try:
                        doc.submit()
                    except Exception:
                        errors.append(f"Failed to submit {doc.name}: {cstr(frappe.get_traceback())[:250]}")

                created.append(doc.name)
                first_to = first_to or doc.name
                per_to.append({"to": doc.name, "items": len(doc.items), "mode": "created"})
                total_items_imported += len(doc.items)

        frappe.db.commit()

        return {
            "ok": True if (created or updated) else False,
            "total_items": total_items_imported,
            "created": created,
            "updated": updated,
            "errors": errors,
            "first_to": first_to,
            "per_to": per_to,
            "sheets_used": {"header": header_sheet, "items": SHEET_ITEM_DEFAULT},
        }

    except Exception as e:
        frappe.db.rollback()
        frappe.log_error(title="WMS Transfer Order Import (Header+Store Columns)", message=frappe.get_traceback())
        return {
            "ok": False,
            "total_items": 0,
            "created": [],
            "updated": [],
            "errors": [cstr(e)],
            "first_to": None,
            "per_to": [],
        }