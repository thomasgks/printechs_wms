# -*- coding: utf-8 -*-
import frappe
from frappe.utils import cint, getdate
import openpyxl


ASN_DT = "WMS ASN"
ASN_ITEM_DT = "WMS ASN Item"


@frappe.whitelist()
def import_wms_asn_excel(file_url: str, submit: int = 0):
    """
    Import WMS ASN + WMS ASN Item from Excel in one action.

    Excel must contain sheets:
      - WMS ASN
      - WMS ASN Item

    "Own name" support:
      - Column 'name' in WMS ASN sheet is used as docname.

    Warehouse mapping supported:
      - default_receiving_warehouse (Warehouse docname)
      - default_receiving_warehouse_code (Warehouse.code)

    Your Excel headers supported:
      - default_receiving_warehouse_name   -> treated as default_receiving_warehouse
      - warehouse_code                     -> treated as default_receiving_warehouse_code
    """

    if not file_url:
        frappe.throw("file_url is required")

    f = frappe.get_doc("File", {"file_url": file_url})
    path = f.get_full_path()

    wb = openpyxl.load_workbook(path, data_only=True)

    if "WMS ASN" not in wb.sheetnames or "WMS ASN Item" not in wb.sheetnames:
        frappe.throw("Excel must contain sheets: WMS ASN, WMS ASN Item")

    ws_asn = wb["WMS ASN"]
    ws_item = wb["WMS ASN Item"]

    asn_rows = list(ws_asn.values)
    item_rows = list(ws_item.values)

    if len(asn_rows) < 2:
        frappe.throw("'WMS ASN' sheet has no data")
    if len(item_rows) < 2:
        frappe.throw("'WMS ASN Item' sheet has no data")

    asn_headers = _norm_headers(asn_rows[0])
    item_headers = _norm_headers(item_rows[0])

    # Build items grouped by parent ASN name
    items_by_parent = {}
    for r in item_rows[1:]:
        if not _has_any_value(r):
            continue

        d = dict(zip(item_headers, r))
        parent = _s(d.get("parent"))
        if not parent:
            continue

        items_by_parent.setdefault(parent, []).append(d)

    created = []
    errors = []

    for r in asn_rows[1:]:
        if not _has_any_value(r):
            continue

        d = dict(zip(asn_headers, r))

        asn_name = _s(d.get("name"))
        supplier = _s(d.get("supplier"))

        if not asn_name:
            errors.append({"name": "(blank)", "error": "Column 'name' is required in WMS ASN sheet (own name mode)"})
            continue
        if not supplier:
            errors.append({"name": asn_name, "error": "Supplier is required"})
            continue

        try:
            if frappe.db.exists(ASN_DT, asn_name):
                errors.append({"name": asn_name, "error": "WMS ASN already exists"})
                continue

            asn = frappe.new_doc(ASN_DT)
            asn.name = asn_name  # own name

            # -----------------------------
            # HEADER FIELD MAPPING
            # -----------------------------
            _set_any(asn, ["supplier"], supplier)

            # Dates
            excel_posting_date = d.get("posting_date") or d.get("shipment_date")
            if excel_posting_date:
                _set_any(asn, ["shipment_date", "posting_date"], getdate(excel_posting_date))

            excel_expected = d.get("expected_arrival_date") or d.get("expected_arrival")
            if excel_expected:
                _set_any(asn, ["expected_arrival", "expected_arrival_date"], getdate(excel_expected))

            # Purchase Order
            _set_any(asn, ["purchase_order", "purchase_order_no", "po"], d.get("purchase_order"))

            # Airway Bill
            _set_any(asn, ["airway_bill", "airway_bill_no"], d.get("airway_bill_no") or d.get("airway_bill"))

            # Shipment Type
            _set_any(asn, ["shipment_type"], d.get("shipment_type"))

            # Currency / rate / totals (if fields exist)
            _set_any(asn, ["currency"], d.get("currency"))
            _set_any(asn, ["conversion_rate"], d.get("conversion_rate"))
            _set_any(asn, ["total_shipped_qty"], d.get("total_shipped_qty"))
            _set_any(asn, ["total_ctn", "total_carton", "total_cartons"], d.get("total_ctn"))

            # Status
            _set_any(asn, ["status"], d.get("status") or "Draft")

            # Title
            _set_any(asn, ["asn_title", "title"], d.get("asn_title"))

            # -----------------------------
            # DEFAULT RECEIVING WAREHOUSE (MANDATORY)
            # Your Excel columns:
            #   - default_receiving_warehouse_name
            #   - warehouse_code
            # -----------------------------
            wh_link = _s(
                d.get("default_receiving_warehouse")
                or d.get("default_receiving_warehouse_name")  # <-- YOUR EXCEL
                or d.get("receiving_warehouse")
                or d.get("receiving_warehouse_name")
            )

            wh_code = _s(
                d.get("default_receiving_warehouse_code")
                or d.get("receiving_warehouse_code")
                or d.get("warehouse_code")  # <-- YOUR EXCEL (maps to Warehouse.code)
            )

            wh_name, wh_code_final = _resolve_warehouse(wh_link=wh_link, wh_code=wh_code)

            if not wh_name or not wh_code_final:
                raise Exception(
                    "Default Receiving Warehouse is mandatory. "
                    "Provide 'default_receiving_warehouse_name' (Warehouse docname) "
                    "or 'warehouse_code' (Warehouse.code) in Excel."
                )

            _set_any(asn, ["default_receiving_warehouse"], wh_name)
            _set_any(asn, ["default_receiving_warehouse_code"], wh_code_final)

            # -----------------------------
            # ITEMS
            # -----------------------------
            item_list = items_by_parent.get(asn_name, [])
            if not item_list:
                errors.append({"name": asn_name, "error": "No items found in 'WMS ASN Item' for this parent"})
                continue

            # ✅ IMPORTANT FIX:
            # Find the correct child table fieldname on WMS ASN which points to "WMS ASN Item"
            item_table_field = _get_child_table_fieldname(asn, ASN_ITEM_DT)
            if not item_table_field:
                raise Exception(
                    f"Cannot find child table field in '{ASN_DT}' pointing to '{ASN_ITEM_DT}'. "
                    f"Open DocType '{ASN_DT}' and confirm the Table field Options = '{ASN_ITEM_DT}'."
                )

            for it in item_list:
                item_code = _s(it.get("item_code"))
                qty = it.get("qty") or it.get("shipped_qty")

                if not item_code:
                    continue
                if qty in (None, ""):
                    continue

                child = asn.append(item_table_field, {})
                _set_any(child, ["item_code"], item_code)
                _set_any(child, ["shipped_qty", "qty"], cint(qty))
                _set_any(child, ["uom"], it.get("uom"))
                _set_any(child, ["carton_id", "carton", "box_id"], it.get("carton_id"))
                _set_any(child, ["po_item_reference", "po_detail", "purchase_order_item"], it.get("po_item_reference"))

            # Insert + optional submit
            asn.insert(ignore_permissions=True)

            if cint(submit):
                asn.submit()

            created.append(asn.name)

        except Exception as e:
            errors.append({"name": asn_name, "error": str(e)})

    frappe.db.commit()
    return {"status": "ok", "count": len(created), "created": created, "errors": errors}


# -----------------------------
# Helpers
# -----------------------------

def _norm_headers(headers):
    out = []
    for h in headers:
        h = "" if h is None else str(h)
        out.append(h.strip().lower().replace(" ", "_"))
    return out


def _has_any_value(row):
    if not row:
        return False
    for v in row:
        if v not in (None, "", " "):
            return True
    return False


def _s(v):
    return ("" if v is None else str(v)).strip()


def _set_any(doc, candidates, value):
    """Set first existing field from candidate list"""
    if value in (None, ""):
        return
    for f in candidates:
        try:
            if doc.meta and doc.meta.has_field(f):
                doc.set(f, value)
                return
        except Exception:
            continue


def _get_child_table_fieldname(parent_doc, child_dt: str) -> str:
    """
    Return the Table fieldname on parent_doc which has options == child_dt.
    If multiple, returns the first match.
    """
    try:
        meta = parent_doc.meta
        for df in meta.fields:
            if not df:
                continue
            if df.fieldtype == "Table" and (df.options or "").strip() == child_dt:
                return df.fieldname

        # fallback: if only one Table field exists, use it
        table_fields = [df.fieldname for df in meta.fields if df and df.fieldtype == "Table"]
        if len(table_fields) == 1:
            return table_fields[0]

    except Exception:
        pass

    return ""


def _resolve_warehouse(wh_link: str = "", wh_code: str = ""):
    """
    Returns (warehouse_name, warehouse_code)
    - wh_link: Warehouse docname (Link)
    - wh_code: Warehouse.code (your custom field 'code')
    """
    wh_name = ""
    wh_code_final = ""

    # If link is provided and exists, take it and fetch code
    if wh_link and frappe.db.exists("Warehouse", wh_link):
        wh_name = wh_link
        wh_code_final = frappe.db.get_value("Warehouse", wh_name, "code") or ""
        return wh_name, wh_code_final

    # If code is provided, find Warehouse by code
    if wh_code:
        wh_name = frappe.db.get_value("Warehouse", {"code": wh_code}, "name") or ""
        if wh_name:
            wh_code_final = wh_code
            return wh_name, wh_code_final

    return "", ""
