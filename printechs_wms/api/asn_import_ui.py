# -*- coding: utf-8 -*-
import frappe
from frappe.utils import cint, getdate
import openpyxl


@frappe.whitelist()
def import_wms_asn_excel(file_url: str, submit: int = 0):
    """
    Import WMS ASN + WMS ASN Item from Excel in one action.

    Excel sheets required:
      - WMS ASN
      - WMS ASN Item

    Supports your template headers (confirmed):
      WMS ASN:
        - default_receiving_warehouse_code
        - default_receiving_warehouse_name
      WMS ASN Item:
        - shipped_qty, received_qty, unit_cost, extended_cost, etc.

    Also supports alternate header:
      - warehouse_code  (as warehouse code)

    Mandatory fields handled (sets whatever exists on your ASN doctype):
      - default_receiving_warehouse / default_receiving_warehouse_name / receiving_warehouse
      - default_receiving_warehouse_code / warehouse_code / receiving_warehouse_code
    """

    if not file_url:
        frappe.throw("file_url is required")

    f = frappe.get_doc("File", {"file_url": file_url})
    path = f.get_full_path()

    wb = openpyxl.load_workbook(path, data_only=True)

    if "WMS ASN" not in wb.sheetnames or "WMS ASN Item" not in wb.sheetnames:
        frappe.throw("Excel must contain sheets: WMS ASN, WMS ASN Item")

    ws_asn = wb["WMS ASN"]
    ws_item = wb["WMS ASN Item"]

    asn_rows = list(ws_asn.values)
    item_rows = list(ws_item.values)

    if len(asn_rows) < 2:
        frappe.throw("'WMS ASN' sheet has no data")
    if len(item_rows) < 2:
        frappe.throw("'WMS ASN Item' sheet has no data")

    asn_headers = _norm_headers(asn_rows[0])
    item_headers = _norm_headers(item_rows[0])

    # Group items by parent ASN
    items_by_parent = {}
    for r in item_rows[1:]:
        if not _has_any_value(r):
            continue
        d = dict(zip(item_headers, r))
        parent = _s(d.get("parent"))
        if not parent:
            continue
        items_by_parent.setdefault(parent, []).append(d)

    created = []
    errors = []

    for r in asn_rows[1:]:
        if not _has_any_value(r):
            continue

        d = dict(zip(asn_headers, r))

        asn_name = _s(d.get("name"))
        supplier = _s(d.get("supplier"))

        if not asn_name:
            errors.append({"name": "(blank)", "error": "Column 'name' is required in WMS ASN sheet"})
            continue
        if not supplier:
            errors.append({"name": asn_name, "error": "Supplier is required"})
            continue

        try:
            if frappe.db.exists("WMS ASN", asn_name):
                errors.append({"name": asn_name, "error": "WMS ASN already exists"})
                continue

            asn = frappe.new_doc("WMS ASN")
            asn.name = asn_name  # own name

            # -----------------------------
            # HEADER FIELD MAPPING
            # -----------------------------
            _set_any(asn, ["supplier"], supplier)

            excel_posting_date = d.get("posting_date") or d.get("shipment_date")
            if excel_posting_date:
                _set_any(asn, ["posting_date", "shipment_date"], getdate(excel_posting_date))

            excel_expected = d.get("expected_arrival_date") or d.get("expected_arrival")
            if excel_expected:
                _set_any(asn, ["expected_arrival_date", "expected_arrival"], getdate(excel_expected))

            _set_any(asn, ["asn_title", "title"], d.get("asn_title"))
            _set_any(asn, ["purchase_order", "purchase_order_no", "po"], d.get("purchase_order"))
            _set_any(asn, ["airway_bill_no", "airway_bill"], d.get("airway_bill_no") or d.get("airway_bill"))
            _set_any(asn, ["shipment_type"], d.get("shipment_type"))

            _set_any(asn, ["currency"], d.get("currency"))
            _set_any(asn, ["conversion_rate"], d.get("conversion_rate"))

            # -----------------------------
            # DEFAULT RECEIVING WAREHOUSE (MANDATORY)
            # template: default_receiving_warehouse_code + default_receiving_warehouse_name
            # -----------------------------
            raw_wh_name = _pick(d,
                "default_receiving_warehouse",
                "default_receiving_warehouse_name",   # TEMPLATE
                "receiving_warehouse",
                "receiving_warehouse_name",
            )

            raw_wh_code = _pick(d,
                "default_receiving_warehouse_code",   # TEMPLATE
                "receiving_warehouse_code",
                "warehouse_code",                     # your earlier file
            )

            wh_name, wh_code = _resolve_warehouse(wh_link=_s(raw_wh_name), wh_code=_s(raw_wh_code))

            if not wh_name or not wh_code:
                raise Exception(
                    "Default Receiving Warehouse is mandatory. "
                    f"Excel provided name='{_s(raw_wh_name)}', code='{_s(raw_wh_code)}'. "
                    f"Resolved name='{wh_name}', code='{wh_code}'. "
                    "Fix Excel value OR ensure Warehouse exists with matching docname/code."
                )

            # Set ALL possible fieldnames if they exist on your doctype
            _set_any(asn, ["default_receiving_warehouse", "receiving_warehouse"], wh_name)
            _set_any(asn, ["default_receiving_warehouse_name"], wh_name)  # in case your fieldname is *_name
            _set_any(asn, ["default_receiving_warehouse_code", "receiving_warehouse_code", "warehouse_code"], wh_code)

            # -----------------------------
            # ITEMS
            # -----------------------------
            item_list = items_by_parent.get(asn_name, [])
            if not item_list:
                errors.append({"name": asn_name, "error": "No items found in 'WMS ASN Item' for this parent"})
                continue

            for it in item_list:
                item_code = _s(it.get("item_code"))
                shipped_qty = it.get("shipped_qty") or it.get("qty")
                received_qty = it.get("received_qty")

                if not item_code:
                    continue
                if shipped_qty in (None, ""):
                    continue

                child = asn.append("items", {})
                _set_any(child, ["item_code"], item_code)
                _set_any(child, ["shipped_qty", "qty"], cint(shipped_qty))
                if received_qty not in (None, ""):
                    _set_any(child, ["received_qty"], cint(received_qty))

                _set_any(child, ["uom"], it.get("uom"))
                _set_any(child, ["carton_id"], it.get("carton_id"))
                _set_any(child, ["unit_cost"], it.get("unit_cost"))
                _set_any(child, ["extended_cost"], it.get("extended_cost"))

            asn.insert(ignore_permissions=True)

            if cint(submit):
                asn.submit()

            created.append(asn.name)

        except Exception as e:
            errors.append({"name": asn_name, "error": str(e)})

    frappe.db.commit()
    return {"status": "ok", "count": len(created), "created": created, "errors": errors}


# -----------------------------
# Helpers
# -----------------------------

def _norm_headers(headers):
    out = []
    for h in headers:
        h = "" if h is None else str(h)
        out.append(h.strip().lower().replace(" ", "_"))
    return out

def _has_any_value(row):
    if not row:
        return False
    for v in row:
        if v not in (None, "", " "):
            return True
    return False

def _s(v):
    return ("" if v is None else str(v)).strip()

def _pick(d: dict, *keys):
    """Return first non-empty value from dict for given keys."""
    for k in keys:
        v = d.get(k)
        if v not in (None, "", " "):
            return v
    return ""

def _set_any(doc, candidates, value):
    """Set first existing field from candidate list."""
    if value in (None, ""):
        return
    for f in candidates:
        try:
            if doc.meta.has_field(f):
                doc.set(f, value)
                return
        except Exception:
            continue

def _resolve_warehouse(wh_link: str = "", wh_code: str = ""):
    """
    Returns (warehouse_docname, warehouse_code)
    - wh_link: Warehouse docname (e.g., 'Main Warehouse - MAATC')
    - wh_code: Warehouse.code (your custom field 'code', e.g., 'WH-MAIN')
    """
    # Prefer link if valid
    if wh_link and frappe.db.exists("Warehouse", wh_link):
        code = frappe.db.get_value("Warehouse", wh_link, "code") or ""
        return wh_link, code

    # Else try code lookup
    if wh_code:
        name = frappe.db.get_value("Warehouse", {"code": wh_code}, "name") or ""
        if name:
            return name, wh_code

    return "", ""
