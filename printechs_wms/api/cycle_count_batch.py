# -*- coding: utf-8 -*-
"""
cycle_count_batch.py

APIs included:
1) sync_task_capture_only(payload)              -> capture task + replace results safely (no duplicates)
2) load_actual_stock_preview(batch_name)        -> compute system_qty/delta + fill Batch Summary
3) export_opening_valuation_template(batch_name)-> export grouped excel for finance to fill valuation_rate
4) upload_opening_valuation_file(...)           -> read excel + create Opening Stock Reconciliation (SR)
5) confirm_and_post_batch(...)                  -> update WMS Stock Balance + link SR (opening) or create SR (adjustment)

Notes:
- FIXED: API4 file path indentation bug (file always resolves)
- FIXED: API4 avoids ERPNext popup "None of the items..." by pre-checking diffs before creating SR
- SAFE: Detects SR item fields dynamically across ERPNext versions
"""

from __future__ import annotations

import frappe
from frappe import _
from frappe.utils import (
    today,
    now_datetime,
    nowdate,
    nowtime,
    cint,
    flt,
    cstr,
)

# ---------------------------------------------------------------------
# DOCTYPES
# ---------------------------------------------------------------------
TASK_DT = "WMS Cycle Count Task"
RESULT_DT = "WMS Cycle Count Result"
BATCH_DT = "WMS Cycle Count Batch"
SUMMARY_CHILD_DT = "WMS Cycle Count Batch Summary"
STOCK_BAL_DT = "WMS Stock Balance"

API_VERSION = "cycle_count_batch_v2_complete"


# ---------------------------------------------------------------------
# CONFIG: Difference Accounts (Update these to your chart of accounts)
# ---------------------------------------------------------------------
OPENING_DIFFERENCE_ACCOUNT = "1.04.01.01 - Temporary Opening - MAATC"     # MUST be Asset/Liability
ADJUSTMENT_DIFFERENCE_ACCOUNT = "5.01.01.01 - Stock Adjustment - MAATC"   # Usually Expense


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
def _get_meta(dt: str):
    return frappe.get_meta(dt)


def _meta_has(dt: str, fieldname: str) -> bool:
    try:
        return _get_meta(dt).has_field(fieldname)
    except Exception:
        return False


def _safe_float(v) -> float:
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _normalize_task_status(in_status: str | None) -> str:
    allowed = {"Open", "In Progress", "Completed", "Validated", "Posted", "Closed", "Cancelled"}
    if not in_status:
        return "Completed"
    s = str(in_status).strip()
    if s in allowed:
        return s
    mapping = {"RECEIVED": "Completed", "DONE": "Completed", "FINISHED": "Completed"}
    return mapping.get(s.upper(), "Completed")


def _task_by_external_ref(external_ref: str):
    if not external_ref:
        return None
    return frappe.db.get_value(TASK_DT, {"external_ref": external_ref}, "name")


def _extract_carton_id(row: dict) -> str | None:
    v = (
        row.get("carton_id")
        or row.get("carton")
        or row.get("ctn_id")
        or row.get("carton_no")
        or row.get("carton_ref")
    )
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def get_or_create_batch(company: str, warehouse: str, warehouse_code: str, posting_date: str | None):
    posting_date = posting_date or today()
    bname = frappe.db.get_value(
        BATCH_DT,
        {
            "company": company,
            "warehouse": warehouse,
            "warehouse_code": warehouse_code,
            "posting_date": posting_date,
            "status": "Draft",
        },
        "name",
    )
    if bname:
        return bname

    b = frappe.get_doc(
        {
            "doctype": BATCH_DT,
            "company": company,
            "warehouse": warehouse,
            "warehouse_code": warehouse_code,
            "posting_date": posting_date,
            "status": "Draft",
        }
    )
    b.insert(ignore_permissions=True)
    return b.name


def link_task_to_batch(task_doc):
    if task_doc.get("batch"):
        return task_doc.batch

    bname = get_or_create_batch(
        task_doc.company,
        task_doc.warehouse,
        task_doc.warehouse_code,
        getattr(task_doc, "posting_date", None),
    )
    task_doc.db_set("batch", bname, update_modified=True)

    if _meta_has(TASK_DT, "sync_stage"):
        task_doc.db_set("sync_stage", "In Batch", update_modified=True)

    return bname


def _pick_difference_account(opening_entry: int) -> str:
    """
    Opening entry requires Asset/Liability account in ERPNext.
    Normal adjustment should go to Expense/Stock Adjustment.
    """
    return OPENING_DIFFERENCE_ACCOUNT if cint(opening_entry) else ADJUSTMENT_DIFFERENCE_ACCOUNT


def _ensure_account_exists(account_name: str):
    if not account_name or not cstr(account_name).strip():
        frappe.throw(_("Difference account is required."))
    if not frappe.db.exists("Account", account_name):
        frappe.throw(_("Account not found: {0}").format(account_name))


def _validate_opening_account(diff_acc: str, company: str):
    """
    Opening SR requires Asset/Liability leaf account for same company.
    """
    _ensure_account_exists(diff_acc)

    acc = frappe.db.get_value(
        "Account",
        diff_acc,
        ["root_type", "company", "is_group", "disabled"],
        as_dict=True,
    )
    if not acc:
        frappe.throw(_("Opening diff account not found: {0}").format(diff_acc))
    if acc.company != company:
        frappe.throw(_("Opening diff account company mismatch: {0}").format(diff_acc))
    if cint(acc.is_group or 0) == 1:
        frappe.throw(_("Opening diff account cannot be group: {0}").format(diff_acc))
    if cint(acc.disabled or 0) == 1:
        frappe.throw(_("Opening diff account is disabled: {0}").format(diff_acc))
    if acc.root_type not in ("Asset", "Liability"):
        frappe.throw(
            _("Opening Entry requires Asset/Liability account. {0} root_type={1}").format(diff_acc, acc.root_type)
        )


def _get_stock_snapshot(item_code: str, warehouse: str, posting_date: str, posting_time: str):
    """
    qty + valuation_rate as of posting date/time (best effort across versions).
    Tries erpnext.stock.utils.get_stock_balance; falls back to Bin (current).
    """
    try:
        from erpnext.stock.utils import get_stock_balance  # type: ignore

        qty, rate = get_stock_balance(
            item_code=item_code,
            warehouse=warehouse,
            posting_date=posting_date,
            posting_time=posting_time,
            with_valuation_rate=True,
        )
        return flt(qty), flt(rate)
    except Exception:
        pass

    b = frappe.db.get_value(
        "Bin",
        {"item_code": item_code, "warehouse": warehouse},
        ["actual_qty", "valuation_rate"],
        as_dict=True,
    ) or {}
    return flt(b.get("actual_qty") or 0), flt(b.get("valuation_rate") or 0)


# ---------------------------------------------------------------------
# API 1: Desktop -> Capture Task (REPLACE results, NO append)
# ---------------------------------------------------------------------
@frappe.whitelist()
def sync_task_capture_only(payload: dict | None = None):
    """
    Desktop pushes task header + lines.
    - If task exists by external_ref -> replace results fully (delete child rows)
    - If not exists -> create task + results
    """
    if payload is None:
        payload = frappe.local.form_dict.get("payload") or frappe.local.form_dict or {}

    task = payload.get("task") or payload.get("header") or payload.get("task_header") or payload
    lines = payload.get("lines") or payload.get("results") or payload.get("items") or []

    if not isinstance(task, dict):
        frappe.throw(_("Invalid payload.task/header (must be dict)."))

    external_ref = (task.get("external_ref") or task.get("external_task_ref") or task.get("task_id") or "").strip()
    if not external_ref:
        frappe.throw(_("external_ref is required (desktop task id)."))

    company = task.get("company")
    warehouse = task.get("warehouse")
    warehouse_code = task.get("warehouse_code")
    bin_location = (task.get("bin_location") or "").strip()

    if not company:
        frappe.throw(_("company is required."))
    if not warehouse:
        frappe.throw(_("warehouse is required (ERP Warehouse link)."))
    if not warehouse_code:
        frappe.throw(_("warehouse_code is required."))

    posting_date = task.get("posting_date") or today()
    status = _normalize_task_status(task.get("status") or "Completed")

    existing_name = _task_by_external_ref(external_ref)

    # Resolve child doctype from task's "results" table (fallback RESULT_DT)
    result_dt = RESULT_DT
    try:
        tf = frappe.get_meta(TASK_DT).get_field("results")
        if tf and getattr(tf, "options", None):
            result_dt = tf.options
    except Exception:
        pass

    result_has_carton = _meta_has(result_dt, "carton_id")

    carton_is_link = False
    carton_reqd = False
    if result_has_carton:
        try:
            f = frappe.get_meta(result_dt).get_field("carton_id")
            if f:
                carton_is_link = (getattr(f, "fieldtype", None) == "Link")
                # if Link -> we do not force set (user asked earlier to use Data)
                if not carton_is_link:
                    carton_reqd = getattr(f, "reqd", False)
        except Exception:
            pass

    def _delete_existing_child_rows(parent_name: str):
        # robust delete using table name
        results_field = frappe.get_meta(TASK_DT).get_field("results")
        if not results_field or not getattr(results_field, "options", None):
            return
        child_doctype = results_field.options

        table = None
        try:
            table = getattr(frappe.get_meta(child_doctype), "table_name", None)
        except Exception:
            table = None

        if not table:
            try:
                table = frappe.db.get_table_name(child_doctype)
            except Exception:
                table = None

        if table:
            frappe.db.sql(f"DELETE FROM `{table}` WHERE parent=%s", (parent_name,))

    if existing_name:
        doc = frappe.get_doc(TASK_DT, existing_name)

        # update header
        doc.company = company
        doc.warehouse = warehouse
        doc.warehouse_code = warehouse_code
        if _meta_has(TASK_DT, "bin_location"):
            doc.bin_location = bin_location
        if _meta_has(TASK_DT, "posting_date"):
            doc.posting_date = posting_date
        if _meta_has(TASK_DT, "status"):
            doc.status = status
        if _meta_has(TASK_DT, "sync_stage"):
            doc.sync_stage = "Captured"
        if _meta_has(TASK_DT, "sync_status"):
            doc.sync_status = "Synced"

        # delete all existing child rows to avoid append duplicates
        _delete_existing_child_rows(doc.name)
        frappe.db.commit()
        doc.reload()
        doc.set("results", [])

        # rebuild results
        if isinstance(lines, list):
            for i, row in enumerate(lines, start=1):
                if not isinstance(row, dict):
                    continue
                item_code = row.get("item_code") or row.get("item") or row.get("code")
                if not item_code:
                    continue
                row_bin = row.get("bin_location") or bin_location
                counted_qty = _safe_float(row.get("counted_qty") or row.get("qty") or row.get("counted") or 0)
                carton_id_val = _extract_carton_id(row)

                if result_has_carton and (not carton_is_link) and carton_reqd:
                    if not (carton_id_val and str(carton_id_val).strip()):
                        frappe.throw(_("Row {0}: carton_id is required (Item {1}).").format(i, item_code))

                child_dict = {
                    "item_code": item_code,
                    "bin_location": row_bin,
                    "counted_qty": counted_qty,
                    "system_qty": 0.0,
                    "delta_qty": counted_qty,
                    "has_discrepancy": 1 if abs(counted_qty) > 0.000001 else 0,
                }
                if result_has_carton and (not carton_is_link):
                    child_dict["carton_id"] = (carton_id_val or "").strip()

                doc.append("results", child_dict)

                # some versions need explicit set
                if result_has_carton and (not carton_is_link) and doc.results:
                    doc.results[-1].carton_id = (carton_id_val or "").strip()

        doc.save(ignore_permissions=True)
        updated_lines = len(doc.results)

    else:
        doc = frappe.get_doc(
            {
                "doctype": TASK_DT,
                "company": company,
                "warehouse": warehouse,
                "warehouse_code": warehouse_code,
                "bin_location": bin_location,
                "status": status,
                "external_ref": external_ref,
                "posting_date": posting_date,
            }
        )
        if _meta_has(TASK_DT, "sync_stage"):
            doc.sync_stage = "Captured"
        if _meta_has(TASK_DT, "sync_status"):
            doc.sync_status = "Synced"

        if isinstance(lines, list):
            for i, row in enumerate(lines, start=1):
                if not isinstance(row, dict):
                    continue
                item_code = row.get("item_code") or row.get("item") or row.get("code")
                if not item_code:
                    continue
                row_bin = row.get("bin_location") or bin_location
                counted_qty = _safe_float(row.get("counted_qty") or row.get("qty") or row.get("counted") or 0)
                carton_id_val = _extract_carton_id(row)

                if result_has_carton and (not carton_is_link) and carton_reqd:
                    if not (carton_id_val and str(carton_id_val).strip()):
                        frappe.throw(_("Row {0}: carton_id is required (Item {1}).").format(i, item_code))

                child_dict = {
                    "item_code": item_code,
                    "bin_location": row_bin,
                    "counted_qty": counted_qty,
                    "system_qty": 0.0,
                    "delta_qty": counted_qty,
                    "has_discrepancy": 1 if abs(counted_qty) > 0.000001 else 0,
                }
                if result_has_carton and (not carton_is_link):
                    child_dict["carton_id"] = (carton_id_val or "").strip()

                doc.append("results", child_dict)

                if result_has_carton and (not carton_is_link) and doc.results:
                    doc.results[-1].carton_id = (carton_id_val or "").strip()

        doc.insert(ignore_permissions=True)
        updated_lines = len(doc.results)

    batch_name = link_task_to_batch(doc)

    return {
        "ok": True,
        "api_version": API_VERSION,
        "task": doc.name,
        "external_ref": external_ref,
        "batch": batch_name,
        "updated_lines": updated_lines,
    }


# ---------------------------------------------------------------------
# API 2: Batch -> Load Actual Stock Preview
# ---------------------------------------------------------------------
@frappe.whitelist()
def load_actual_stock_preview(batch_name: str):
    b = frappe.get_doc(BATCH_DT, batch_name)

    warehouse_code = b.warehouse_code
    if not warehouse_code:
        frappe.throw(_("Batch.warehouse_code is required."))

    task_names = frappe.get_all(TASK_DT, filters={"batch": batch_name}, pluck="name")
    if not task_names:
        return {"ok": True, "updated_lines": 0, "summary_rows": 0, "note": "No tasks linked to this batch"}

    result_has_carton = _meta_has(RESULT_DT, "carton_id")

    summary_carton_reqd = False
    if _meta_has(SUMMARY_CHILD_DT, "carton_id"):
        try:
            sf = frappe.get_meta(SUMMARY_CHILD_DT).get_field("carton_id")
            if sf:
                summary_carton_reqd = getattr(sf, "reqd", False)
        except Exception:
            pass

    res_fields = ["name", "parent", "item_code", "bin_location", "counted_qty"]
    if result_has_carton:
        res_fields.append("carton_id")

    res_rows = frappe.get_all(
        RESULT_DT,
        filters={"parent": ["in", task_names]},
        fields=res_fields,
        limit_page_length=200000,
    )

    def _k(r):
        if result_has_carton:
            return (r["item_code"], r["bin_location"], (r.get("carton_id") or ""))
        return (r["item_code"], r["bin_location"])

    # One counted_qty per key -> keep latest row by name (prevents triple totals)
    grouped = {}
    for r in res_rows:
        key = _k(r)
        q = _safe_float(r.get("counted_qty"))
        rname = r.get("name") or ""
        if key not in grouped or (rname > (grouped[key].get("name") or "")):
            grouped[key] = {"counted": q, "name": rname}

    b.set("summary", [])
    system_map = {}

    for key, agg in grouped.items():
        if result_has_carton:
            item_code, bin_location, carton_id = key
        else:
            item_code, bin_location = key
            carton_id = None

        where = ["warehouse = %s", "location = %s", "item_code = %s"]
        params = [b.warehouse, bin_location, item_code]

        if carton_id and _meta_has(STOCK_BAL_DT, "carton"):
            where.append("carton = %s")
            params.append(carton_id)

        system_qty = frappe.db.sql(
            f"""
            SELECT COALESCE(SUM(qty), 0)
            FROM `tab{STOCK_BAL_DT}`
            WHERE {" AND ".join(where)}
            """,
            tuple(params),
        )[0][0] or 0

        system_qty = float(system_qty)
        counted = float(agg["counted"])
        delta = counted - system_qty

        system_map[key] = system_qty

        summary_row = {
            "item_code": item_code,
            "bin_location": bin_location,
            "total_system_qty": system_qty,
            "total_counted_qty": counted,
            "total_delta_qty": delta,
        }

        if _meta_has(SUMMARY_CHILD_DT, "carton_id"):
            summary_row["carton_id"] = (carton_id or "").strip() if result_has_carton else ""
            if summary_row["carton_id"] == "" and summary_carton_reqd:
                summary_row["carton_id"] = "-"  # placeholder if mandatory

        b.append("summary", summary_row)

    # update each result row system_qty/delta
    updated_lines = 0
    for r in res_rows:
        key = _k(r)
        sys_qty = float(system_map.get(key, 0.0))
        counted = _safe_float(r.get("counted_qty"))
        delta = counted - sys_qty
        has_disc = 1 if abs(delta) > 0.000001 else 0

        frappe.db.set_value(RESULT_DT, r["name"], "system_qty", sys_qty, update_modified=False)
        frappe.db.set_value(RESULT_DT, r["name"], "delta_qty", delta, update_modified=False)
        frappe.db.set_value(RESULT_DT, r["name"], "has_discrepancy", has_disc, update_modified=False)
        updated_lines += 1

    if _meta_has(BATCH_DT, "preview_loaded_on"):
        b.preview_loaded_on = now_datetime()

    if _meta_has(BATCH_DT, "status"):
        b.status = "Previewed"

    b.save(ignore_permissions=True)

    return {"ok": True, "batch": batch_name, "updated_lines": updated_lines, "summary_rows": len(b.summary)}


# ---------------------------------------------------------------------
# API 3: Export Opening Valuation Template (Excel)
# ---------------------------------------------------------------------
@frappe.whitelist()
def export_opening_valuation_template(batch_name=None):
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from frappe.utils.file_manager import save_file

    if not batch_name:
        batch_name = frappe.local.form_dict.get("batch_name")
    if not batch_name:
        frappe.throw(_("batch_name is required"))

    b = frappe.get_doc(BATCH_DT, batch_name)
    company = b.get("company")
    warehouse = b.get("warehouse")
    posting_date = b.get("posting_date")

    if not company:
        frappe.throw(_("Batch.company is required"))
    if not warehouse:
        frappe.throw(_("Batch.warehouse is required"))

    task_names = frappe.get_all(TASK_DT, filters={"batch": batch_name}, pluck="name")
    if not task_names:
        frappe.throw(_("No tasks linked to this batch"))

    result_meta = frappe.get_meta(RESULT_DT)

    fields = ["parent", "item_code", "counted_qty"]
    if result_meta.has_field("bin_location"):
        fields.append("bin_location")
    if result_meta.has_field("location"):
        fields.append("location")
    if result_meta.has_field("carton_id"):
        fields.append("carton_id")
    if result_meta.has_field("carton"):
        fields.append("carton")

    dedupe_uuid_field = None
    for cand in ["event_uuid", "offline_uuid", "sync_uuid", "uuid"]:
        if result_meta.has_field(cand):
            dedupe_uuid_field = cand
            fields.append(cand)
            break

    res_rows = frappe.get_all(
        RESULT_DT,
        filters={
            "parent": ["in", task_names],
            "parenttype": TASK_DT,
        },
        fields=fields,
        limit_page_length=200000,
    )

    if not res_rows:
        frappe.throw(_("No result lines found for tasks in this batch"))

    # 1) Dedupe by UUID (best)
    if dedupe_uuid_field:
        seen = set()
        clean_rows = []
        for r in res_rows:
            uid = cstr(r.get(dedupe_uuid_field)).strip()
            if not uid:
                clean_rows.append(r)
                continue
            if uid in seen:
                continue
            seen.add(uid)
            clean_rows.append(r)
        res_rows = clean_rows

    # 2) Else fallback dedupe by (task,item,loc,carton) -> keep MAX qty
    else:
        loc_field = "bin_location" if result_meta.has_field("bin_location") else ("location" if result_meta.has_field("location") else None)
        carton_field = "carton_id" if result_meta.has_field("carton_id") else ("carton" if result_meta.has_field("carton") else None)

        best = {}
        for r in res_rows:
            parent = cstr(r.get("parent")).strip()
            item_code = cstr(r.get("item_code")).strip()
            if not item_code:
                continue
            loc = cstr(r.get(loc_field)).strip() if loc_field else ""
            carton = cstr(r.get(carton_field)).strip() if carton_field else ""
            key = (parent, item_code, loc, carton)

            qty = flt(r.get("counted_qty") or 0)
            if key not in best or qty > best[key]:
                best[key] = qty

        res_rows = [{"item_code": k[1], "counted_qty": v} for k, v in best.items()]

    # group by item_code
    grouped = {}
    for r in res_rows:
        item_code = cstr(r.get("item_code")).strip()
        if not item_code:
            continue
        grouped[item_code] = grouped.get(item_code, 0.0) + flt(r.get("counted_qty") or 0)

    if not grouped:
        frappe.throw(_("No item totals to export"))

    item_codes = list(grouped.keys())
    item_info = {
        x["name"]: x
        for x in frappe.get_all(
            "Item",
            filters={"name": ["in", item_codes]},
            fields=["name", "item_name", "stock_uom"],
        )
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opening Valuation Upload"

    headers = [
        "company",
        "warehouse",
        "posting_date",
        "item_code",
        "item_name",
        "uom",
        "counted_qty_total",
        "valuation_rate",
        "currency",
        "remarks",
    ]
    ws.append(headers)

    for item_code in sorted(item_codes):
        inf = item_info.get(item_code, {})
        ws.append(
            [
                company,
                warehouse,
                str(posting_date) if posting_date else "",
                item_code,
                inf.get("item_name") or "",
                inf.get("stock_uom") or "",
                grouped.get(item_code, 0.0),
                "",  # finance fills
                "SAR",
                f"Batch {batch_name}: fill valuation_rate for opening stock items",
            ]
        )

    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5597")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(14, min(30, len(h) + 4))

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("README")
    readme = [
        "Finance fills valuation_rate for opening stock items.",
        "Upload this file using API 4: upload_opening_valuation_file.",
        "Then run API 5 confirm_and_post_batch with is_opening=1.",
        "",
        "Difference accounts (update in code if needed):",
        f"- Opening: {OPENING_DIFFERENCE_ACCOUNT}",
        f"- Adjustment: {ADJUSTMENT_DIFFERENCE_ACCOUNT}",
        "",
        "NOTE: Export dedupes Result lines to avoid double/triple qty caused by sync retries.",
    ]
    for i, line in enumerate(readme, start=1):
        ws2.cell(row=i, column=1, value=line).alignment = Alignment(wrap_text=True)
    ws2.column_dimensions["A"].width = 120

    buff = io.BytesIO()
    wb.save(buff)
    data = buff.getvalue()

    filename = f"Opening-Valuation-{batch_name}.xlsx"
    f = save_file(filename, data, BATCH_DT, batch_name, is_private=1)

    return {
        "ok": True,
        "api_version": API_VERSION,
        "batch": batch_name,
        "file_name": f.file_name,
        "file_url": f.file_url,
        "item_count": len(item_codes),
        "dedupe_by": dedupe_uuid_field or "parent+item+location+carton(MAX)",
    }


# ---------------------------------------------------------------------
# API 4: Upload Valuation Excel -> Create Opening Stock SR
# ---------------------------------------------------------------------
# ---------------------------------------------------------------------
# API 4: Upload Valuation Excel -> Create Opening Stock SR (FIXED)
# ---------------------------------------------------------------------
@frappe.whitelist()
def upload_opening_valuation_file(file_url=None, file_id=None, batch_name=None, submit=1):
    """
    Reads sheet "Opening Valuation Upload" and creates Stock Reconciliation (Opening Stock).

    Fixes:
    1) FIXED INDENTATION: file resolve runs even if batch_name is not provided
    2) CRITICAL: DOES NOT call sr.insert() before adding items (prevents EmptyStockReconciliationItemsError)
    3) Pre-check diffs (best effort) and returns friendly response if no change
    4) Detects SR item qty/rate field names dynamically across ERPNext versions
    """
    import openpyxl
    from frappe.utils import cint, nowdate, flt, cstr

    submit = cint(submit or 1)

    # ---- security (finance only) ----
    allowed_roles = {"Accounts Manager", "Stock Manager", "System Manager"}
    if not any(r in allowed_roles for r in frappe.get_roles(frappe.session.user)):
        frappe.throw(_("Not allowed. Finance/Stock Manager only."))

    # ---- read params from request if missing ----
    if not file_url and not file_id:
        file_url = frappe.local.form_dict.get("file_url")
        file_id = frappe.local.form_dict.get("file_id")
    if not batch_name:
        batch_name = frappe.local.form_dict.get("batch_name")

    if not file_url and not file_id:
        frappe.throw(_("file_url or file_id is required"))

    # ---- batch defaults ----
    batch_defaults = {}
    b = None
    user_diff_account = (
        frappe.local.form_dict.get("difference_account")
        or frappe.local.form_dict.get("opening_difference_account")
    )

    if batch_name:
        b = frappe.get_doc(BATCH_DT, batch_name)
        batch_defaults = {
            "company": b.get("company"),
            "warehouse": b.get("warehouse"),
            "posting_date": b.get("posting_date"),
        }

        # prevent duplicate opening SR
        if _meta_has(BATCH_DT, "opening_stock_reconciliation"):
            existing_sr = (b.get("opening_stock_reconciliation") or "").strip()
            if existing_sr and frappe.db.exists("Stock Reconciliation", existing_sr):
                frappe.throw(
                    _(
                        "Opening Stock Reconciliation already done for this batch: {0}. "
                        "Cancel that SR first if you need to create a new one."
                    ).format(existing_sr)
                )

        if not user_diff_account and _meta_has(BATCH_DT, "difference_account"):
            user_diff_account = (b.get("difference_account") or "").strip() or None
        if not user_diff_account and _meta_has(BATCH_DT, "opening_difference_account"):
            user_diff_account = (b.get("opening_difference_account") or "").strip() or None

    # -----------------------------------------------------------------
    # FIXED: Resolve file path (NOT inside `if batch_name`)
    # -----------------------------------------------------------------
    fdoc = None
    if file_id:
        fdoc = frappe.get_doc("File", file_id)
    else:
        file_url = cstr(file_url).strip()
        if not file_url:
            frappe.throw(_("file_url or file_id is required"))
        file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
        if not file_name:
            frappe.throw(_("File not found for URL: {0}").format(file_url))
        fdoc = frappe.get_doc("File", file_name)

    if not fdoc:
        frappe.throw(_("Could not resolve file. Provide file_id or file_url."))

    file_path = fdoc.get_full_path()

    # -----------------------------------------------------------------
    # Read Excel
    # -----------------------------------------------------------------
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = "Opening Valuation Upload"
    if sheet_name not in wb.sheetnames:
        frappe.throw(_("Sheet '{0}' not found").format(sheet_name))
    ws = wb[sheet_name]

    header = [(c.value or "").strip() if isinstance(c.value, str) else (c.value or "") for c in ws[1]]

    def idx(col):
        try:
            return header.index(col)
        except ValueError:
            return None

    required_cols = ["company", "warehouse", "posting_date", "item_code", "counted_qty_total", "valuation_rate"]
    for col in required_cols:
        if idx(col) is None:
            frappe.throw(_("Missing column in Excel: {0}").format(col))

    parsed = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(vals):
            continue

        company = vals[idx("company")] or batch_defaults.get("company")
        warehouse = vals[idx("warehouse")] or batch_defaults.get("warehouse")
        posting_date = vals[idx("posting_date")] or batch_defaults.get("posting_date") or nowdate()

        item_code = vals[idx("item_code")]
        if isinstance(item_code, str):
            item_code = item_code.strip()

        counted_qty = vals[idx("counted_qty_total")] or 0
        valuation_rate = vals[idx("valuation_rate")] or 0

        if not company or not warehouse:
            frappe.throw(_("Row {0}: company/warehouse is required (or provide batch_name)").format(r))
        if not item_code:
            frappe.throw(_("Row {0}: item_code is required").format(r))

        try:
            counted_qty = float(counted_qty or 0)
        except Exception:
            frappe.throw(_("Row {0}: counted_qty_total must be a number").format(r))

        try:
            valuation_rate = float(valuation_rate or 0)
        except Exception:
            frappe.throw(_("Row {0}: valuation_rate must be a number").format(r))

        if counted_qty < 0:
            frappe.throw(_("Row {0}: counted_qty_total cannot be negative").format(r))
        if valuation_rate <= 0:
            frappe.throw(_("Row {0}: valuation_rate is required (> 0)").format(r))

        if not frappe.db.exists("Item", item_code):
            frappe.throw(_("Row {0}: Item not found: {1}").format(r, item_code))

        parsed.append(
            {
                "company": str(company).strip(),
                "warehouse": str(warehouse).strip(),
                "posting_date": posting_date,
                "item_code": str(item_code).strip(),
                "qty": counted_qty,
                "valuation_rate": valuation_rate,
            }
        )

    if not parsed:
        frappe.throw(_("No valid rows found in Excel"))

    first_company = parsed[0]["company"]
    first_wh = parsed[0]["warehouse"]
    first_posting_date = parsed[0]["posting_date"] or nowdate()
    posting_time = "00:00:00"  # opening safety

    # Safety: only one company+warehouse in file
    for x in parsed:
        if x["company"] != first_company or x["warehouse"] != first_wh:
            frappe.throw(_("Excel must contain one company + one warehouse only (for safety)."))

    # -----------------------------------------------------------------
    # Aggregate by item_code (sum qty, weighted avg rate)
    # -----------------------------------------------------------------
    agg = {}
    for x in parsed:
        item = x["item_code"]
        q = float(x["qty"] or 0)
        v = float(x["valuation_rate"] or 0)
        if item not in agg:
            agg[item] = {"qty": 0.0, "total_val": 0.0}
        agg[item]["qty"] += q
        agg[item]["total_val"] += q * v

    rows = []
    for item, a in agg.items():
        qty = float(a["qty"] or 0)
        rate = (float(a["total_val"]) / qty) if qty else 0.0
        rows.append({"item_code": item, "qty": qty, "valuation_rate": rate})

    # -----------------------------------------------------------------
    # Difference Account (Opening must be Asset/Liability)
    # -----------------------------------------------------------------
    diff_acc = (cstr(user_diff_account).strip() or "") or _pick_difference_account(opening_entry=1)
    _ensure_account_exists(diff_acc)

    acc = frappe.db.get_value("Account", diff_acc, ["root_type", "company", "is_group", "disabled"], as_dict=True)
    if not acc:
        frappe.throw(_("Opening diff account not found: {0}").format(diff_acc))
    if acc.company != first_company:
        frappe.throw(_("Opening diff account company mismatch: {0}").format(diff_acc))
    if cint(acc.is_group or 0) == 1:
        frappe.throw(_("Opening diff account cannot be group: {0}").format(diff_acc))
    if cint(acc.disabled or 0) == 1:
        frappe.throw(_("Opening diff account is disabled: {0}").format(diff_acc))
    if acc.root_type not in ("Asset", "Liability"):
        frappe.throw(_("Opening Entry requires Asset/Liability account. {0} root_type={1}")
                    .format(diff_acc, acc.root_type))

    # -----------------------------------------------------------------
    # Detect SR item fields (ERPNext version safe)
    # -----------------------------------------------------------------
    SR_DT = "Stock Reconciliation"
    SR_ITEM_DT = "Stock Reconciliation Item"
    sr_meta = frappe.get_meta(SR_DT)
    sri_meta = frappe.get_meta(SR_ITEM_DT)

    qty_field = "qty" if sri_meta.has_field("qty") else ("quantity" if sri_meta.has_field("quantity") else None)
    if not qty_field:
        frappe.throw(_("SR Item has no qty/quantity field."))

    val_field = "valuation_rate" if sri_meta.has_field("valuation_rate") else ("rate" if sri_meta.has_field("rate") else None)
    if not val_field:
        frappe.throw(_("SR Item has no valuation_rate field."))

    # -----------------------------------------------------------------
    # OPTIONAL: Pre-check diffs (best effort)
    # -----------------------------------------------------------------
    def _snapshot(item_code: str):
        # use helper from your file if exists; else fallback Bin
        try:
            return _get_stock_snapshot(item_code, first_wh, str(first_posting_date), posting_time)
        except Exception:
            b = frappe.db.get_value(
                "Bin",
                {"item_code": item_code, "warehouse": first_wh},
                ["actual_qty", "valuation_rate"],
                as_dict=True,
            ) or {}
            return flt(b.get("actual_qty") or 0), flt(b.get("valuation_rate") or 0)

    filtered = []
    no_change = []
    for x in sorted(rows, key=lambda z: z["item_code"]):
        cur_qty, cur_rate = _snapshot(x["item_code"])
        new_qty = flt(x["qty"])
        new_rate = flt(x["valuation_rate"])
        if abs(new_qty - cur_qty) < 1e-9 and abs(new_rate - cur_rate) < 1e-9:
            no_change.append(
                {
                    "item_code": x["item_code"],
                    "current_qty": cur_qty,
                    "new_qty": new_qty,
                    "current_rate": cur_rate,
                    "new_rate": new_rate,
                }
            )
        else:
            filtered.append(x)

    if not filtered:
        return {
            "ok": False,
            "api_version": API_VERSION,
            "message": "No change detected for all items. Stock Reconciliation not created.",
            "company": first_company,
            "warehouse": first_wh,
            "posting_date": first_posting_date,
            "difference_account": diff_acc,
            "no_change_count": len(no_change),
            "sample_no_change": no_change[:10],
        }

    rows = filtered

    # -----------------------------------------------------------------
    # CRITICAL FIX: Create SR WITH ITEMS then insert
    # -----------------------------------------------------------------
    sr = frappe.get_doc({"doctype": SR_DT})

    if sr_meta.has_field("company"):
        sr.company = first_company
    if sr_meta.has_field("posting_date"):
        sr.posting_date = first_posting_date
    if sr_meta.has_field("posting_time"):
        sr.posting_time = posting_time
    if sr_meta.has_field("purpose"):
        sr.purpose = "Opening Stock"
    if sr_meta.has_field("opening_entry"):
        sr.opening_entry = 1
    if sr_meta.has_field("set_warehouse"):
        sr.set_warehouse = first_wh
    if sr_meta.has_field("expense_account"):
        sr.expense_account = diff_acc
    if sr_meta.has_field("difference_account"):
        sr.difference_account = diff_acc

    sr.set("items", [])

    seen = set()
    for x in sorted(rows, key=lambda z: z["item_code"]):
        key = (x["item_code"], first_wh)
        if key in seen:
            frappe.throw(_("Duplicate item+warehouse detected: Item {0} Warehouse {1}").format(x["item_code"], first_wh))
        seen.add(key)

        sr.append(
            "items",
            {
                "item_code": x["item_code"],
                "warehouse": first_wh,
                qty_field: float(x["qty"]),
                val_field: float(x["valuation_rate"]),
            },
        )

    # Insert AFTER items exist (so validate doesn't see empty items)
    sr.insert(ignore_permissions=True)

    if submit:
        sr.submit()

    # Link SR to batch
    if batch_name and _meta_has(BATCH_DT, "opening_stock_reconciliation"):
        b = frappe.get_doc(BATCH_DT, batch_name)
        b.opening_stock_reconciliation = sr.name
        b.save(ignore_permissions=True)

    return {
        "ok": True,
        "api_version": API_VERSION,
        "sr": sr.name,
        "docstatus": sr.docstatus,
        "company": first_company,
        "warehouse": first_wh,
        "difference_account": diff_acc,
        "row_count_excel": len(parsed),
        "row_count_after_agg": len(agg),
        "row_count_in_sr": len(rows),
        "skipped_no_change_count": len(no_change),
        "skipped_sample": no_change[:10],
    }


# ---------------------------------------------------------------------
# API 5: Confirm & Post Batch
# ---------------------------------------------------------------------
@frappe.whitelist()
def confirm_and_post_batch(batch_name=None, create_stock_reconciliation=1, is_opening=0):
    """
    FULL UPDATED API 5 (NO DUPLICATE SR + NO POPUP ERROR + POSTS ALWAYS)

    Key fixes:
    1) If Batch already Posted -> return immediately (prevents double click / retry duplicates)
    2) Opening mode (is_opening=1) -> NEVER create SR here (uses API4 SR)
    3) Adjustment mode (is_opening=0):
       - Reuse existing Batch.stock_reconciliation if already set
       - When creating SR, catch ERPNext "None of the items have any change..." and SKIP SR (no throw)
    """
    create_stock_reconciliation = cint(create_stock_reconciliation or 0)
    is_opening = cint(is_opening or 0)

    if not batch_name:
        batch_name = frappe.local.form_dict.get("batch_name")
    if not batch_name:
        frappe.throw(_("batch_name is required"))

    b = frappe.get_doc(BATCH_DT, batch_name)
    status = (b.get("status") or "").strip()

    # ------------------------------------------------------------------
    # 0) If already posted -> stop here (prevents duplicate SR on re-click)
    # ------------------------------------------------------------------
    if status == "Posted":
        existing_sr = (b.get("stock_reconciliation") or "").strip() if _meta_has(BATCH_DT, "stock_reconciliation") else ""
        existing_sr = existing_sr or None
        return {
            "ok": True,
            "api_version": API_VERSION,
            "batch": batch_name,
            "mode": "opening" if cint(is_opening) else "adjustment",
            "message": "Batch already posted. No action taken.",
            "updated_balances": 0,
            "sr": existing_sr,
        }

    if status not in {"Draft", "Previewed", ""}:
        frappe.throw(_("Batch is not in allowed status. Current status: {0}").format(status))

    company = (b.get("company") or "").strip()
    warehouse = (b.get("warehouse") or "").strip()
    posting_date = b.get("posting_date") or nowdate()
    posting_time = nowtime()

    if not company:
        frappe.throw(_("Batch.company is required"))
    if not warehouse:
        frappe.throw(_("Batch.warehouse is required"))

    # Opening SR must exist (created by API 4)
    opening_sr_field = "opening_stock_reconciliation"
    opening_sr_name = b.get(opening_sr_field) if _meta_has(BATCH_DT, opening_sr_field) else None
    if is_opening and not opening_sr_name:
        frappe.throw(_("Opening mode: Upload valuation Excel and create Opening Stock SR first (API 4)."))

    nowdt = now_datetime()

    # ------------------------------------------------------------------
    # 1) Load tasks + results
    # ------------------------------------------------------------------
    task_names = frappe.get_all(TASK_DT, filters={"batch": batch_name}, pluck="name")
    if not task_names:
        frappe.throw(_("No tasks linked to this batch."))

    has_carton_id = _meta_has(RESULT_DT, "carton_id")
    res_fields = ["item_code", "bin_location", "counted_qty"]
    if has_carton_id:
        res_fields.append("carton_id")

    res_rows = frappe.get_all(
        RESULT_DT,
        filters={"parent": ["in", task_names]},
        fields=res_fields,
        limit_page_length=200000,
    )
    if not res_rows:
        frappe.throw(_("No result lines found for tasks in this batch."))

    # ------------------------------------------------------------------
    # 2) Update WMS Stock Balance by (item, location, carton)
    # ------------------------------------------------------------------
    wms_group = {}
    for r in res_rows:
        item_code = (r.get("item_code") or "").strip()
        location = (r.get("bin_location") or "").strip()
        carton = ((r.get("carton_id") if has_carton_id else None) or "").strip()
        qty = float(r.get("counted_qty") or 0)

        if not item_code or not location:
            continue

        key = (item_code, location, carton)
        wms_group[key] = wms_group.get(key, 0.0) + qty

    updated_balances = 0
    for (item_code, location, carton_key), counted_qty in wms_group.items():
        carton = carton_key or None

        if not frappe.db.exists("Item", item_code):
            frappe.throw(_("Item not found: {0}").format(item_code))

        filters = {"company": company, "warehouse": warehouse, "item_code": item_code, "location": location}
        if carton:
            filters["carton"] = carton
        else:
            filters["carton"] = ["in", ["", None]]

        existing = frappe.db.get_value(STOCK_BAL_DT, filters, "name")

        if existing:
            frappe.db.set_value(STOCK_BAL_DT, existing, "qty", float(counted_qty), update_modified=False)
            if _meta_has(STOCK_BAL_DT, "last_txn_datetime"):
                frappe.db.set_value(STOCK_BAL_DT, existing, "last_txn_datetime", nowdt, update_modified=False)
        else:
            bal = frappe.get_doc(
                {
                    "doctype": STOCK_BAL_DT,
                    "company": company,
                    "warehouse": warehouse,
                    "item_code": item_code,
                    "location": location,
                    "carton": carton,
                    "qty": float(counted_qty),
                }
            )
            if _meta_has(STOCK_BAL_DT, "last_txn_datetime"):
                bal.last_txn_datetime = nowdt
            bal.insert(ignore_permissions=True)

        updated_balances += 1

    # ------------------------------------------------------------------
    # 3) Stock Reconciliation (NO DUPLICATES + NO "NO CHANGE" POPUP)
    # ------------------------------------------------------------------
    sr_name = None
    sr_note = None

    existing_batch_sr = None
    if _meta_has(BATCH_DT, "stock_reconciliation"):
        existing_batch_sr = (b.get("stock_reconciliation") or "").strip() or None
        if existing_batch_sr and not frappe.db.exists("Stock Reconciliation", existing_batch_sr):
            existing_batch_sr = None

    if is_opening:
        # Opening: SR comes from API 4 only
        sr_name = opening_sr_name
        sr_note = "Opening mode: SR not created here; linked API4 Opening SR."

        # optional: store it for easy reference in UI (do not override)
        if _meta_has(BATCH_DT, "stock_reconciliation") and not existing_batch_sr and sr_name:
            b.stock_reconciliation = sr_name

    else:
        # Adjustment:
        if existing_batch_sr:
            sr_name = existing_batch_sr
            sr_note = "Adjustment SR already linked on batch; reused (no duplicate)."
        else:
            if create_stock_reconciliation:
                # group by item_code
                erp_group = {}
                for r in res_rows:
                    item_code = (r.get("item_code") or "").strip()
                    qty = float(r.get("counted_qty") or 0)
                    if not item_code:
                        continue
                    erp_group[item_code] = erp_group.get(item_code, 0.0) + qty

                SR_DT = "Stock Reconciliation"
                SR_ITEM_DT = "Stock Reconciliation Item"
                sr_meta = frappe.get_meta(SR_DT)
                sri_meta = frappe.get_meta(SR_ITEM_DT)

                qty_field = "qty" if sri_meta.has_field("qty") else ("quantity" if sri_meta.has_field("quantity") else None)
                if not qty_field:
                    frappe.throw(_("Stock Reconciliation Item missing qty field"))

                val_field = "valuation_rate" if sri_meta.has_field("valuation_rate") else ("rate" if sri_meta.has_field("rate") else None)

                diff_acc = _pick_difference_account(opening_entry=0)
                _ensure_account_exists(diff_acc)

                sr = frappe.get_doc({"doctype": SR_DT})

                if sr_meta.has_field("company"):
                    sr.company = company
                if sr_meta.has_field("posting_date"):
                    sr.posting_date = posting_date
                if sr_meta.has_field("posting_time"):
                    sr.posting_time = posting_time

                if sr_meta.has_field("purpose"):
                    df = sr_meta.get_field("purpose")
                    opts = [x.strip() for x in (df.options or "").split("\n") if x.strip()]
                    preferred = "Stock Reconciliation"
                    sr.purpose = preferred if preferred in opts else (opts[0] if opts else preferred)

                if sr_meta.has_field("set_warehouse"):
                    sr.set_warehouse = warehouse

                if sr_meta.has_field("expense_account"):
                    sr.expense_account = diff_acc
                if sr_meta.has_field("difference_account"):
                    sr.difference_account = diff_acc

                sr.set("items", [])

                for item_code, counted_qty in erp_group.items():
                    bin_row = frappe.db.get_value(
                        "Bin",
                        {"item_code": item_code, "warehouse": warehouse},
                        ["valuation_rate"],
                        as_dict=True,
                    )
                    vr = float((bin_row or {}).get("valuation_rate") or 0)

                    if vr <= 0:
                        frappe.throw(
                            _(
                                "Valuation missing for Item {0} in Warehouse {1}. "
                                "If this is Opening Stock, run API 4 then call Confirm with is_opening=1."
                            ).format(item_code, warehouse)
                        )

                    row = {"item_code": item_code, "warehouse": warehouse, qty_field: float(counted_qty)}
                    if val_field:
                        row[val_field] = vr
                    sr.append("items", row)

                # IMPORTANT: catch ERPNext "no change" error and SKIP SR
                try:
                    sr.insert(ignore_permissions=True)
                    sr.submit()
                    sr_name = sr.name
                    sr_note = "Adjustment SR created."

                    if _meta_has(BATCH_DT, "stock_reconciliation"):
                        b.stock_reconciliation = sr_name

                except Exception as e:
                    msg = str(e) or ""
                    if "None of the items have any change in quantity or value" in msg or "None of the items have any change" in msg:
                        # Do not throw -> no popup, batch can still be posted
                        sr_name = None
                        sr_note = "Adjustment mode: ERPNext detected no change; SR skipped."
                    else:
                        raise

            else:
                sr_note = "Adjustment mode: create_stock_reconciliation=0; SR not created."

    # ------------------------------------------------------------------
    # 4) Post the batch + tasks
    # ------------------------------------------------------------------
    if _meta_has(BATCH_DT, "status"):
        b.status = "Posted"
    if _meta_has(BATCH_DT, "posted_on"):
        b.posted_on = nowdt
    if _meta_has(BATCH_DT, "posted_by"):
        b.posted_by = frappe.session.user

    b.save(ignore_permissions=True)

    for tname in task_names:
        if _meta_has(TASK_DT, "status"):
            frappe.db.set_value(TASK_DT, tname, "status", "Posted", update_modified=False)
        if _meta_has(TASK_DT, "sync_stage"):
            frappe.db.set_value(TASK_DT, tname, "sync_stage", "Posted", update_modified=False)

    return {
        "ok": True,
        "api_version": API_VERSION,
        "batch": batch_name,
        "mode": "opening" if is_opening else "adjustment",
        "updated_balances": updated_balances,
        "sr": sr_name,
        "sr_note": sr_note,
        "difference_account_opening": OPENING_DIFFERENCE_ACCOUNT,
        "difference_account_adjustment": ADJUSTMENT_DIFFERENCE_ACCOUNT,
    }